import weatherhat
import time
import csv
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Initialiseer de sensor
sensor = weatherhat.WeatherHAT()

try: #probeer excel file in te laden
 workbook = load_workbook('weather_data.xlsx')
 sheet = workbook.active
except FileNotFoundError: #als file niet bestaat maak hem aan 
 workbook = Workbook()
 sheet = workbook.active
 sheet.append(["Tijdstip", "Apparaat Temperatuur (°C)", "Temperatuur (°C)",
                     "Luchtdruk (hPa)", "Luchtvochtigheid (%)", "Lichtintensiteit (lux)", 
                     "Windsnelheid (m/s)", "Windrichting (graden)"])


while True:
    try:
        # Ververs de sensorwaarden
        sensor.update()
        
        # Lees sensorwaarden
        apparaat_temperatuur = sensor.device_temperature
        temperatuur = sensor.temperature
        druk = sensor.pressure
        vochtigheid = sensor.humidity
        lux = sensor.lux
        windkracht = sensor.wind_speed
        windrichting = sensor.wind_direction

        # Print de waarden in de console
        tijdstip = datetime.now().strftime('%d-%m-%y %H:%M:%S')
        print(f"Tijdstip: {tijdstip}")
        print(f"Apparaat temperatuur: {apparaat_temperatuur:.2f} °C")
        print(f"Temperatuur: {temperatuur:.2f} °C")
        print(f"Luchtdruk: {druk:.2f} hPa")
        print(f"Luchtvochtigheid: {vochtigheid:.2f} %")
        print(f"Lichtintensiteit: {lux:.2f} lux")
        print(f"Windsnelheid: {windkracht:.2f} m/s")
        print(f"Windrichting: {windrichting:.2f} graden")
            
        sheet.append([tijdstip, apparaat_temperatuur, temperatuur, 
                     druk, vochtigheid, lux, windkracht, windrichting])
        workbook.save('weather_data.xlsx')

        # Wacht 10 seconden
        time.sleep(10)

    except Exception as e:
        fout_tijdstip = datetime.now().strftime('%d-%m-%y %H:%M:%S')
        print(f"Fout opgetreden op {fout_tijdstip}: {e}")
        time.sleep(5)  # Wacht even voordat je opnieuw probeert
