import time
import pygame
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Pygame-instellingen
pygame.init()
width, height = 800, 600
window = pygame.display.set_mode((width, height))
pygame.display.set_caption("Weergave Namen en Tijd")
font = pygame.font.Font(None, 30)

# Excel-bestand laden of aanmaken
try:  # Probeer de Excel-bestand in te laden
    workbook = load_workbook('snelle_data.xlsx')
    sheet = workbook.active
except FileNotFoundError:  # Als bestand niet bestaat, maak hem aan
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Tijdstip", "Naam1", "Naam2", "Naam3", "Naam4", "Naam5", "Naam6", "Naam7"])

# Namen instellen
Naam1 = "henk"
Naam2 = "slim"
Naam3 = "shrek"
Naam4 = "alberto"
Naam5 = "sendy"
Naam6 = "dave"
Naam7 = "frodo"

# Functie om een grafiek te tekenen en op te slaan als afbeelding
def update_line_chart():
    tijden = []
    naam1_data = []
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tijden.append(row[0])  # Tijdstippen
        naam1_data.append(len(row[1]))  # Data voor Naam1 als voorbeeld
        
    plt.figure(figsize=(6, 4))
    plt.plot(tijden, naam1_data, label="Naam1 lengte")
    plt.xlabel("Tijdstip")
    plt.ylabel("Lengte van Naam1")
    plt.title("Voorbeeld Lijndiagram")
    plt.legend()
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig("line_chart.png")
    plt.close()

# Hoofdprogramma-lus
running = True
refresh_time = 3  # ververstijd in seconden
last_update = time.time() - refresh_time  # zorgt dat de eerste keer direct wordt uitgevoerd

while running:
    # Event-lus om de app interactief te houden
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False
    
    # Ververs de tijd en namen in het venster als de refreshtijd verstreken is
    if time.time() - last_update >= refresh_time:
        last_update = time.time()
        tijdstip = datetime.now().strftime('%d-%m-%y %H:%M:%S')
        
        # Schrijf naar Excel
        sheet.append([tijdstip, Naam1, Naam2, Naam3, Naam4, Naam5, Naam6, Naam7])
        workbook.save('snelle_data.xlsx')
        
        # Update en laad de grafiek als afbeelding
        update_line_chart()
        chart_image = pygame.image.load("line_chart.png")
        chart_rect = chart_image.get_rect(center=(width - 300, height // 2))
        
        # Venster vernieuwen
        window.fill((255, 255, 255))  # Witte achtergrond
        tijd_text = font.render(f"Tijdstip: {tijdstip}", True, (0, 0, 0))
        naam1_text = font.render(f"Naam1: {Naam1}", True, (0, 0, 0))
        
        # Teksten en grafiek weergeven
        window.blit(tijd_text, (10, 10))
        window.blit(naam1_text, (10, 70))
        window.blit(chart_image, chart_rect.topleft)  # Grafiek weergave

        pygame.display.flip()

# Pygame afsluiten
pygame.quit()
