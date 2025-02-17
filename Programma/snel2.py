import time
import pygame
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Pygame-instellingen
pygame.init()
width, height = 600, 500
window = pygame.display.set_mode((width, height))
pygame.display.set_caption("Weergave Namen en Tijd")
font = pygame.font.Font(None, 30)

# Excel-bestand laden of aanmaken
try:  # Probeer de Excel-bestand in te laden
    workbook = load_workbook('snelle_data.xlsx')
    sheet = workbook.active
except FileNotFoundError:  # Als bestand niet bestaat, maak hem aan
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Tijdstip", "Naam1", "Naam2", "Naam3", "Naam4", "Naam5", "Naam6", "Naam7"])

# Namen instellen
Naam1 = "henk"
Naam2 = "slim"
Naam3 = "shrek"
Naam4 = "alberto"
Naam5 = "sendy"
Naam6 = "dave"
Naam7 = "frodo"

# Hoofdprogramma-lus
running = True
refresh_time = 3  # ververstijd in seconden
last_update = time.time() - refresh_time  # zorgt dat de eerste keer direct wordt uitgevoerd

while running:
    # Event-lus om de app interactief te houden
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False
    
    # Ververs de tijd en namen in het venster als de refreshtijd verstreken is
    if time.time() - last_update >= refresh_time:
        last_update = time.time()
        tijdstip = datetime.now().strftime('%d-%m-%y %H:%M:%S')
        
        # Schrijf naar Excel
        sheet.append([tijdstip, Naam1, Naam2, Naam3, Naam4, Naam5, Naam6, Naam7])
        workbook.save('snelle_data.xlsx')
        
        # Venster vernieuwen
        window.fill((255, 255, 255))  # Witte achtergrond
        tijd_text = font.render(f"Tijdstip: {tijdstip}", True, (0, 0, 0))
        naam1_text = font.render(f"Naam1: {Naam1}", True, (0, 0, 0))
        naam2_text = font.render(f"Naam2: {Naam2}", True, (0, 0, 0))
        naam3_text = font.render(f"Naam3: {Naam3}", True, (0, 0, 0))
        naam4_text = font.render(f"Naam4: {Naam4}", True, (0, 0, 0))
        naam5_text = font.render(f"Naam5: {Naam5}", True, (0, 0, 0))
        naam6_text = font.render(f"Naam6: {Naam6}", True, (0, 0, 0))
        naam7_text = font.render(f"Naam7: {Naam7}", True, (0, 0, 0))
        
        
         # Teksten weergeven
        window.blit(tijd_text, (10, 10))
        window.blit(naam1_text, (10, 70))
        window.blit(naam2_text, (10, 130))
        window.blit(naam3_text, (10, 190))
        window.blit(naam4_text, (10, 250))
        window.blit(naam5_text, (10, 310))
        window.blit(naam6_text, (10, 370))
        window.blit(naam7_text, (10, 430))
        print(Naam1)
        print(Naam2)
        print(Naam3)
        print(Naam4)
        print(Naam5)
        print(Naam6)
        print(Naam7)
        
        pygame.display.flip()
        
        pygame.display.update()

# Pygame afsluiten
pygame.quit()
